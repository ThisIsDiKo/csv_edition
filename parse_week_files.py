from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd

wb = load_workbook('./2018 180D1.xlsx')
print(wb.sheetnames)
sheet = wb['TDSheet']
print(sheet.title)

print(sheet.values)
data = sheet.values
columns = next(data)[1:]
print(columns)
d = list(data)
for i in d:
    print(i)